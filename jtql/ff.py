#!/user/bin/env python
# -*- coding: UTF-8 -*-
import os
import openpyxl

def getNum(_str):
    if _str is None:
        return 0
    else:
        return float(_str)

filesLocation = 'D:\\excel'

# 列
columnList = [8, 9]
# 页
sheetList = ['2018年度', '2019年度', '2020年度', '2021年1至5月底']
# 部门
fileList = []
for root, dirs, files in os.walk(filesLocation):
    for file in files:
        fileList.append(os.path.join(root, file).split("\\")[2].split(".")[0])
departmentList = fileList
print('部门\t和\tH38\tH43\tH45')
for department_id in range(len(departmentList)):
    workbook = openpyxl.load_workbook(filesLocation + '/' + departmentList[department_id] + '.xlsx')
    worksheet = workbook[sheetList[3]]
    v1 = getNum(worksheet.cell(row=12, column=8).value)
    v2 = getNum(worksheet.cell(row=13, column=8).value)
    v3 = getNum(worksheet.cell(row=14, column=8).value)
    v4 = getNum(worksheet.cell(row=15, column=8).value)
    v = v1 + v2 + v3 + v4
    b = getNum(worksheet.cell(row=38, column=8).value)
    n = getNum(worksheet.cell(row=43, column=8).value)
    m = getNum(worksheet.cell(row=45, column=8).value)
    print(str(departmentList[department_id]) + '\t' + str(v) + '\t' + str(b) + '\t' + str(n) + '\t' + str(m))