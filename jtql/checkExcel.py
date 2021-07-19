#!/user/bin/env python
# -*- coding: UTF-8 -*-
import os
import openpyxl

# 部门
departmentList = []
filesLocation = 'D:\\excel'
for root, dirs, files in os.walk(filesLocation):
    for file in files:
        departmentList.append(os.path.join(root, file).split("\\")[2].split(".")[0])

# 页
sheetList = ['2017年度', '2018年度', '2019年度', '2020年度', '2021年1至5月底']

# 行
rowList = [12, 13, 14, 15, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 38, 39, 40, 43, 44, 45, 46,
           47, 48, 51, 52, 53, 54, 55, 56, 57, 60, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 77, 78, 79, 80, 81, 82, 85,
           86, 87, 88, 89, 90, 91, 92, 93, 94, 97, 98, 99, 100, 101, 104, 105, 106, 107, 108, 109, 112, 113]
# 列
# 8H：发放人数
# 9I：年人均发放标准
# 14N：资金规模
# 15O：个税
# 16P：住房公积金
# 17Q：缴纳养老保险情况
columnList = [8, 9]

for department_id in range(len(departmentList)):
    for sheet_id in range(len(sheetList)):
        workbook = openpyxl.load_workbook(filesLocation + "\\" + departmentList[department_id] + '.xlsx')
        print(filesLocation + "\\" + departmentList[department_id] + '.xlsx')
        worksheet = workbook[sheetList[sheet_id]]
        for row_id in range(len(rowList)):
            for column_id in range(len(columnList)):
                value = worksheet.cell(row=rowList[row_id], column=columnList[column_id]).value
                if value is None:
                    pass
                else:
                    if str(value).find('元') >= 0 or str(value).find('人') >= 0 or str(value).find('=') >= 0:
                        info = departmentList[department_id] + "-->" + sheetList[sheet_id] + \
                               "-->单元格行号:" + str(rowList[row_id]) + "，列号:" + str(columnList[column_id]) + "数据异常！值为：" + value
                        print(info)
        value1 = worksheet.cell(row=7, column=15).value
        value2 = worksheet.cell(row=7, column=16).value
        value3 = worksheet.cell(row=7, column=17).value
        if value1 is None:
            info = departmentList[department_id] + "-->" + sheetList[sheet_id] + \
                   "-->缴纳个人所得税情况数据异常！值为：" + str(value1)
            print(info)
        if value2 is None:
            info = departmentList[department_id] + "-->" + sheetList[sheet_id] + \
                   "-->缴存住房公积金情况数据异常！值为：" + str(value2)
            print(info)
        if value3 is None:
            info = departmentList[department_id] + "-->" + sheetList[sheet_id] + \
                   "-->缴纳养老保险情况数据异常！值为：" + str(value3)
            print(info)
        vv = worksheet.cell(row=113, column=1).value
        if vv != '……':
            info = departmentList[department_id] + "-->" + sheetList[sheet_id] + \
                   "-->该页面可能被添加额外的行"
            print(info)
