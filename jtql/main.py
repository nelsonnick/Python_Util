#!/usr/bin/python
# -*- coding: UTF-8 -*-
import openpyxl
def getNum(_str):
    if _str is None:
        return 0
    else:
        return float(_str)
def round_up(value):
    return round(value * 100) / 100.0
# 行
rowList = [12, 13, 14, 15, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 38, 39, 40, 43, 44, 45, 46,
           47, 48, 51, 52, 53, 54, 55, 56, 57, 60, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 77, 78, 79, 80, 81, 82, 85,
           86, 87, 88, 89, 90, 91, 92, 93, 94, 97, 98, 99, 100, 101, 104, 105, 106, 107, 108, 109, 112, 113]
#
columnList = [8, 14]
bumen = ['编办',
         '党校',
         '机关工委',
         '纪委',
         '科协',
         '农业农村局',
         '区体育事业发展中心',
         '区委办',
         '区政协',
         '人社局',
         '审计局',
         '水务局',
         '统计局',
         '投促局',
         '团区委',
         '退役军人局',
         '文旅局',
         '行政审批服务局',
         '宣传部',
         '医保局',
         '政法委',
         '玉清湖',
         '振兴街',
         '张庄路'
         ]
jiedao = ['玉清湖',
          '振兴街',
          '张庄路']
test = ['行政审批服务局']
# departmentList = ['纪委监委', '区委办', '组织部', '宣传部', '统战部', '政法委', '区委编办', '机关工委', '老干部局', '党校', '总工会', '团委','妇联','科协','侨联','残联','工商联','人大','政协','法院','检察院','政府办','发改局','教体局','科技局','工信局','民政局','司法局','财政局','人社局','自然资源局','住建局','城管局','水务局','农业农村局','商务局','文旅局','卫健局','退役军人局','应急管理局','审计局','市场监管局','统计局','医保局','投促局','行政审批局','信访局','经济开发区','体育事业发展中心','振兴街','五里沟','青年公园','南辛庄','中大槐树','道德街','营市街','西市场','张庄路','段店北路','匡山','美里湖','兴福','玉清湖','腊山','吴家堡']
departmentList = test
sheetList = ['2018年度', '2019年度', '2020年度', '2021年1至5月底']
# 叠加数据
c = [[[[0 for row in range(0, len(rowList))] for column in range(0, len(columnList))] for n in range(0, len(departmentList))]
     for m in range(0, len(sheetList))]
for n in range(0, len(departmentList)):
    workbook = openpyxl.load_workbook('d://excel/' + departmentList[n] + '.xlsx')
    for m in range(0, len(sheetList)):
        worksheet = workbook[sheetList[m]]
        for i in range(0, len(rowList)):
            for j in range(0, len(columnList)):
                # print(str(departmentList[n])+','+str(sheetList[m])+','+str(rowList[i])+','+str(columnList[j]))
                # print(worksheet.cell(row=rowList[i], column=columnList[j]).value)
                c[m][n][j][i] = getNum(worksheet.cell(row=rowList[i], column=columnList[j]).value)

# 乘积数据
t = [[[0 for i in range(0, len(rowList))] for n in range(0, len(departmentList))] for m in range(0, len(sheetList))]

for n in range(0, len(departmentList)):
    workbook = openpyxl.load_workbook('d://excel/' + departmentList[n] + '.xlsx')
    for m in range(0, len(sheetList)):
        worksheet = workbook[sheetList[m]]
        for i in range(0, len(rowList)):
            t[m][n][i] = c[m][n][i][8] * c[m][n][i][9]


# 叠加数据求和
workbook = openpyxl.load_workbook('d://excel/汇总.xlsx')
for m in range(0, len(sheetList)):
    for i in range(0, len(rowList)):
        for j in range(0, len(columnList)):
            worksheet = workbook[sheetList[m]]
            num = 0
            for n in range(0, len(departmentList)):
                num = num + c[m][n][j][i]
            worksheet.cell(rowList[i], columnList[j]).value = round_up(num)

            sumTotal = 0
            for n in range(0, len(departmentList)):
                sumTotal = sumTotal + t[m][n][i]
            worksheet.cell(rowList[i], 19).value = round_up(sumTotal)
workbook.save('d://excel/汇总.xlsx')

# 总计求和
rowList2 = [7]
columnList2 = [15, 16, 17]
s = [[[[0 for i in range(0, len(rowList2))] for j in range(0, len(columnList2))] for n in range(0, len(departmentList))]
     for m in range(0, len(sheetList))]
for n in range(0, len(departmentList)):
    workbook = openpyxl.load_workbook('d://excel/' + departmentList[n] + '.xlsx')
    for m in range(0, len(sheetList)):
        worksheet = workbook[sheetList[m]]
        for i in range(0, len(rowList2)):
            for j in range(0, len(columnList2)):
                s[m][n][j][i] = getNum(worksheet.cell(row=rowList2[i], column=columnList2[j]).value)

workbook = openpyxl.load_workbook('d://excel/汇总.xlsx')
for m in range(0, len(sheetList)):
    for i in range(0, len(rowList2)):
        for j in range(0, len(columnList2)):
            worksheet = workbook[sheetList[m]]
            num = 0
            for n in range(0, len(departmentList)):
                num = num + s[m][n][j][i]
            worksheet.cell(rowList2[i], columnList2[j]).value = round_up(num)
    print('已完成：' + sheetList[m])
workbook.save('d://excel/汇总.xlsx')

print('汇总完成')
