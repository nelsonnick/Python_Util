#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import openpyxl


def getNum(_str):
    if _str is None:
        return 0
    else:
        return float(_str)


def round_up(value):
    return round(value * 100) / 100.0


# name = '综合'
name = '部门'
# name = '街道'
saveLocation = 'C:\\Users\\admin\\Desktop\\报表0706\\' + name + '.xlsx'
filesLocation = 'C:\\Users\\admin\\Desktop\\报表0706\\' + name
# 行
rowList = [12, 13, 14, 15, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 38, 39, 40, 43, 44, 45, 46,
           47, 48, 51, 52, 53, 54, 55, 56, 57, 60, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 77, 78, 79, 80, 81, 82, 85,
           86, 87, 88, 89, 90, 91, 92, 93, 94, 97, 98, 99, 100, 101, 104, 105, 106, 107, 108, 109, 112, 113]
# 列
columnList = [8, 9]
# 页
sheetList = ['2017年度', '2018年度', '2019年度', '2020年度', '2021年1至5月底']
# 部门
fileList = []
for root, dirs, files in os.walk(filesLocation):
    for file in files:
        fileList.append(os.path.join(root, file).split("\\")[6].split(".")[0])

departmentList = fileList
# 第8、14列叠加
c = [[[[0 for column in range(len(columnList))] for row in range(len(rowList))] for sheet in
      range(len(sheetList))]
     for department in range(len(departmentList))]
for department_id in range(len(departmentList)):
    workbook = openpyxl.load_workbook(filesLocation + '/' + departmentList[department_id] + '.xlsx')
    for sheet_id in range(len(sheetList)):
        worksheet = workbook[sheetList[sheet_id]]
        for row_id in range(len(rowList)):
            for column_id in range(len(columnList)):
                c[department_id][sheet_id][row_id][column_id] = getNum(
                    worksheet.cell(row=rowList[row_id], column=columnList[column_id]).value)

workbook = openpyxl.load_workbook(saveLocation)
for sheet_id in range(len(sheetList)):
    for column_id in range(len(columnList)):
        for row_id in range(len(rowList)):
            worksheet = workbook[sheetList[sheet_id]]
            num = 0
            for department_id in range(len(departmentList)):
                num = num + c[department_id][sheet_id][row_id][column_id]
            worksheet.cell(rowList[row_id], columnList[column_id]).value = round_up(num)
workbook.save(saveLocation)
# 平均发放水平
t = [[[0 for row in range(len(rowList))] for sheet in range(len(sheetList))] for department in
     range(len(departmentList))]
for department_id in range(len(departmentList)):
    workbook = openpyxl.load_workbook(filesLocation + '/' + departmentList[department_id] + '.xlsx')
    for sheet_id in range(len(sheetList)):
        worksheet = workbook[sheetList[sheet_id]]
        for row_id in range(len(rowList)):
            t[department_id][sheet_id][row_id] = c[department_id][sheet_id][row_id][0] * \
                                                 c[department_id][sheet_id][row_id][1]

workbook = openpyxl.load_workbook(saveLocation)
for sheet_id in range(len(sheetList)):
    for column_id in range(0, len(columnList)):
        for row_id in range(0, len(rowList)):
            worksheet = workbook[sheetList[sheet_id]]
            num = 0
            for department_id in range(len(departmentList)):
                num = num + t[department_id][sheet_id][row_id]
                if worksheet.cell(rowList[row_id], columnList[0]).value != 0:
                    worksheet.cell(rowList[row_id], columnList[1]).value = num / worksheet.cell(rowList[row_id],
                                                                                                columnList[0]).value
                else:
                    worksheet.cell(rowList[row_id], columnList[1]).value = 0
workbook.save(saveLocation)

# 三项的数组
colList = [15, 16, 17]
r = [[[0 for col in range(len(colList))] for sheet in range(len(sheetList))] for department in
     range(len(departmentList))]
for department_id in range(len(departmentList)):
    workbook = openpyxl.load_workbook(filesLocation + '/' + departmentList[department_id] + '.xlsx')
    for sheet_id in range(len(sheetList)):
        worksheet = workbook[sheetList[sheet_id]]
        for col_id in range(len(colList)):
            r[department_id][sheet_id][col_id] = getNum(worksheet.cell(7, colList[col_id]).value)

# 人数合计的数组
p = [[0 for sheet in range(len(sheetList))] for department in range(len(departmentList))]
for department_id in range(len(departmentList)):
    for sheet_id in range(len(sheetList)):
        p[department_id][sheet_id] = c[department_id][sheet_id][0][0] + c[department_id][sheet_id][1][0] + \
                                     c[department_id][sheet_id][2][0] + c[department_id][sheet_id][3][0]

# 乘积和的数组
x = [[[0 for col in range(len(colList))] for sheet in range(len(sheetList))] for department in
     range(len(departmentList))]
for department_id in range(len(departmentList)):
    for sheet_id in range(len(sheetList)):
        for col_id in range(len(colList)):
            # try:
                x[department_id][sheet_id][col_id] = r[department_id][sheet_id][col_id] * p[department_id][sheet_id]
            # except TypeError:
            #     print(departmentList[department_id])

workbook = openpyxl.load_workbook(saveLocation)


for sheet_id in range(len(sheetList)):
    num0 = 0
    num1 = 0
    num2 = 0
    worksheet = workbook[sheetList[sheet_id]]
    for department_id in range(len(departmentList)):
        num0 = num0 + x[department_id][sheet_id][0]
        num1 = num1 + x[department_id][sheet_id][1]
        num2 = num2 + x[department_id][sheet_id][2]
        personTotal = getNum(worksheet.cell(12, 8).value) + getNum(worksheet.cell(13, 8).value) + \
                      getNum(worksheet.cell(14, 8).value) + getNum(worksheet.cell(15, 8).value)
        worksheet.cell(7, colList[0]).value = num0 / personTotal
        worksheet.cell(7, colList[1]).value = num1 / personTotal
        worksheet.cell(7, colList[2]).value = num2 / personTotal

workbook.save(saveLocation)


