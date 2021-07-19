#!/user/bin/env python
# -*- coding: UTF-8 -*-
import os
import openpyxl


def copy(row_new, row_old):
    worksheet1.cell(row=row_new, column=8).value = worksheet0.cell(row=row_old, column=8).value
    worksheet1.cell(row=row_new, column=9).value = worksheet0.cell(row=row_old, column=9).value
    worksheet1.cell(row=row_new, column=14).value = (worksheet0.cell(row=row_old, column=8).value * worksheet0.cell(row=row_old, column=9).value) / 10000


def setGongHui():
    worksheet1.cell(row=106, column=8).value = worksheet0.cell(row=104, column=8).value
    p1 = worksheet0.cell(row=104, column=8).value
    m1 = worksheet0.cell(row=104, column=9).value
    p2 = worksheet0.cell(row=105, column=8).value
    m2 = worksheet0.cell(row=105, column=9).value
    p3 = worksheet0.cell(row=106, column=8).value
    m3 = worksheet0.cell(row=106, column=9).value
    p4 = worksheet0.cell(row=107, column=8).value
    m4 = worksheet0.cell(row=107, column=9).value
    p5 = worksheet0.cell(row=108, column=8).value
    m5 = worksheet0.cell(row=108, column=9).value
    worksheet1.cell(row=106, column=9).value = (p1*m1 + p2*m2 + p3*m3 + p4*m4 + p5*m5)/p1
    worksheet1.cell(row=106, column=14).value = (p1*m1 + p2*m2 + p3*m3 + p4*m4 + p5*m5)/10000



filename = '街道'
# 列
columnList = [8, 9]
# 页
sheetList = ['2017年度', '2018年度', '2019年度', '2020年度', '2021年1至5月底']

filesLocation0 = 'C:\\Users\\admin\\Desktop\\报表0706'
filesLocation1 = 'C:\\Users\\admin\\Desktop\\报表0716'

workbook0 = openpyxl.load_workbook(filesLocation0 + '/' + filename + '.xlsx')
workbook1 = openpyxl.load_workbook(filesLocation1 + '/' + filename + '.xlsx')

for sheet_id in range(len(sheetList)):
    worksheet0 = workbook0[sheetList[sheet_id]]
    worksheet1 = workbook1[sheetList[sheet_id]]
    worksheet1.cell(row=7, column=15).value = worksheet0.cell(row=7, column=15).value
    worksheet1.cell(row=7, column=16).value = worksheet0.cell(row=7, column=16).value
    worksheet1.cell(row=7, column=17).value = worksheet0.cell(row=7, column=17).value

    # 新/原
    copy(10, 12)    # 职务（职级）工资
    copy(11, 13)    # 级别工资
    copy(16, 20)    # 人民警察警衔津贴
    copy(17, 21)    # 人民警察法定工作日之外加班补贴
    copy(18, 22)    # 人民警察值勤岗位津贴
    copy(19, 23)    # 信访岗位津贴
    copy(20, 24)    # 纪检监察办案人员补贴
    copy(21, 25)    # 审计人员工作补贴
    copy(22, 26)    # 密码人员岗位津贴
    copy(23, 27)    # 税务征收津贴
    copy(24, 28)    # 安全生产监管监察岗位津贴
    copy(25, 29)    # 政法委机关工作津贴
    copy(26, 30)    # 司法助理员岗位津贴
    copy(28, 31)    # 法官、检察官工改保留津贴
    copy(30, 32)    # 转隶保留津贴补贴
    copy(27, 33)    # 乡镇工作补贴

    copy(37, 38)    # 规范津贴补贴（工作性津贴+生活性补贴）

    copy(39, 43)    # 住房补贴
    copy(41, 44)    # 住宅采暖补贴
    copy(42, 45)    # 物业服务补贴
    copy(43, 46)    # 公务交通补贴

    copy(45, 52)    # 精神文明奖
    copy(56, 51)    # 年终一次性奖金
    copy(59, 53)    # 公务员一次性奖金

    copy(62, 66)    # 省市选派第一书记工作补助
    copy(63, 67)    # 东西部协作重庆干部生活补助
    copy(64, 68)    # 东西部协作甘肃干部生活补助
    copy(65, 69)    # “千名（万名）干部下基层”服务队工作补助
    copy(66, 70)    # 市县选派“加强农村基层党组织建设”工作补助
    copy(67, 71)    # 省派基层工作组工作补助

    copy(85, 89)    # 幼儿园收费报销
    copy(86, 90)    # 幼机关食堂补助
    copy(87, 91)    # 加班餐报销
    copy(88, 93)    # 回族等禁猪少数民族伙食补助

    copy(101, 97)    # 应休未休年休假工资报酬
    copy(103, 99)    # 独生子女父母奖励费
    # copy(48, 98)    # 经济社会发展综合考核奖/绩效考核物质奖励

    copy(143, 14)    # 员额法官检察官职务等级工资

    setGongHui()   # 工会

workbook1.save(filesLocation1 + '/' + filename + '.xlsx')
