#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import openpyxl

filesLocation = 'D:\\excel3'

rowList = [35, 40, 48, 57, 60, 74, 82, 94, 101, 109, 113]
columnList = [1]
sheetList = ['2018年度', '2019年度', '2020年度', '2021年1至5月底']

fileList = []
for root, dirs, files in os.walk(filesLocation):
    for file in files:
        fileList.append(os.path.join(root, file).split("\\")[2].split(".")[0])
departmentList = fileList
# ……
# 112   1、
# 56   6、
# 39   2、
# 34   15、
for department_id in range(0, len(departmentList)):
    workbook = openpyxl.load_workbook(filesLocation + '/' + departmentList[department_id] + '.xlsx')
    for sheet_id in range(0, len(sheetList)):
        worksheet = workbook[sheetList[sheet_id]]
        for row_id in range(0, len(rowList)):
            for column_id in range(0, len(columnList)):
                vv = worksheet.cell(row=rowList[row_id], column=columnList[column_id]).value
                if vv != '……':
                    info = str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]) + \
                           "-->第" + str(rowList[row_id]) + "行有新增！原值为：……，新值：" + str(vv)
                    print(info)
        v1 = worksheet.cell(row=34, column=1).value.strip()
        if v1 != '15、':
            info = str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]) + \
                   "-->第34行有新增！新增：" + v1
            print(info)
        v2 = worksheet.cell(row=39, column=1).value.strip()
        if v2 != '2、':
            info = str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]) + \
                   "-->第39行有新增！新增：" + v2
            print(info)
        # print(str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]))
        v3 = worksheet.cell(row=56, column=1).value.strip()
        if v3 != '6、':
            info = str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]) + \
                   "-->第56行有新增！新增：" + v3
            print(info)
        v4 = worksheet.cell(row=112, column=1).value.strip()
        if v4 != '1、':
            info = str(departmentList[department_id]) + "-->" + str(sheetList[sheet_id]) + \
                   "-->第112行有新增！新增：" + v4
            print(info)
