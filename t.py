#!/user/bin/env python
# -*- coding: UTF-8 -*-
import os
import openpyxl

departmentList = []
for root, dirs, files in os.walk(r"D:\excel"):
    for file in files:
        departmentList.append(os.path.join(root, file).split("\\")[2].split(".")[0])

sheetList = []

for i in range(0, len(departmentList)):
    for j in range(0, len(sheetList)):
        workbook = openpyxl.load_workbook("D:\\excel\\" + departmentList[i] + '.xlsx')
        worksheet = workbook.get_sheet_by_name(sheetList[j])
        f = open("D:\\txt\\" + departmentList[i] + '.txt', "wb+")
        if worksheet.cell(row=2, column=3).value > 10 \
                or worksheet.cell(row=2, column=3).value < 5 \
                or worksheet.cell(row=2, column=3).value == 0:
            info = departmentList[i] + "-->" + sheetList[j] + \
                   "-->单元格XX数据异常！正常值：XX；当前值：" + \
                   worksheet.cell(row=2, column=3).value
            f.write(info.encode())
            f.write('\n')
            print(info)

        f.close()
